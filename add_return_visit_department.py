# -*- coding: utf-8 -*-

# 添加访问部门

# 导入处理Excel文件的库
import os
import sys
import json
import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook

import tkinter as tk
from tkinter import messagebox, filedialog

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

config_file = 'config.json'
DEFAULT_SAVE_PATH = r'E:\共享文件\\4杜兴海\\6、综治信息平台纠纷附件\\2026矛盾纠纷附件'

DEPARTMENTS = {
    "1575067821891678306": "金都社区",
    "1575067812324470842": "旗洋社区",
    "1575067825221959730": "大碑社区",
    "1575067817936449608": "水景湾社区",
    "1575067799875776539": "珉谷社区",
    "1575067810499952641": "纳尧村",
    "1575067802501414912": "顶肖村",
    "1575067806737662014": "河堡村"
}

def load_config():
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_config(config):
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def get_base_url():
    """从配置文件获取服务器基础URL"""
    config = load_config()
    host = config.get('server_host', 'your_server_ip_here')
    port = config.get('server_port', '39999')
    return f'https://{host}:{port}'

def select_file_and_token():
    config = load_config()
    current_default = config.get('default_save_path', DEFAULT_SAVE_PATH)
    current_token = config.get('token', '')
    
    root = tk.Tk()
    root.title("选择Excel文件和输入令牌")
    root.geometry("700x250")
    root.resizable(False, False)
    
    selected_file = tk.StringVar()
    token_var = tk.StringVar(value=current_token)
    save_token_var = tk.BooleanVar(value=False)
    
    result = {'file_path': '', 'token': current_token, 'cancelled': True, 'save_token': False}
    
    def browse_file():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            selected_file.set(file_path)
    
    def confirm():
        result['file_path'] = selected_file.get()
        result['token'] = token_var.get()
        result['cancelled'] = False
        result['save_token'] = save_token_var.get()
        root.destroy()
    
    def cancel():
        result['cancelled'] = True
        root.destroy()
    
    def on_closing():
        result['cancelled'] = True
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Excel文件选择
    file_frame = tk.Frame(root)
    file_frame.pack(pady=15, fill=tk.X, padx=20)
    tk.Label(file_frame, text="请选择Excel文件：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    entry_frame = tk.Frame(file_frame)
    entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(entry_frame, textvariable=selected_file, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(entry_frame, text="浏览...", command=browse_file, relief=tk.FLAT, bg="#f0f0f0", padx=10, pady=2).pack(side=tk.LEFT)
    
    # Token输入框
    token_frame = tk.Frame(root)
    token_frame.pack(pady=15, fill=tk.X, padx=20)
    tk.Label(token_frame, text="用户令牌：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    token_entry_frame = tk.Frame(token_frame)
    token_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(token_entry_frame, textvariable=token_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Checkbutton(token_entry_frame, text="保存令牌", variable=save_token_var).pack(side=tk.LEFT, padx=10)
    
    # 按钮区域
    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)
    tk.Button(button_frame, text="确定", command=confirm, width=12, bg="#4CAF50", fg="white", relief=tk.FLAT, padx=15, pady=5).pack(side=tk.LEFT, padx=10)
    tk.Button(button_frame, text="取消", command=cancel, width=12, relief=tk.FLAT, bg="#f0f0f0", padx=15, pady=5).pack(side=tk.LEFT, padx=10)
    
    # 设置默认按钮为回车键
    root.bind('<Return>', lambda event: confirm())
    
    # 居中显示窗口
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()
    return result

def get_department_id(department_name):
    for dept_id, dept_name in DEPARTMENTS.items():
        if dept_name == department_name:
            return dept_id
    return None

def add_return_visit(token, issue_id, circulation_id, department_name):
    url = f"{get_base_url()}/event-center/api/circulation/returnVisit"
    
    visit_dept_id = get_department_id(department_name)
    if not visit_dept_id:
        return {"success": False, "message": f"未找到部门ID: {department_name}"}
    
    opinion = f"请{department_name}及时对该矛盾纠纷案件开展回访工作，做好情况了解与记录。"
    
    data = {
        "issueId": issue_id,
        "circulationId": circulation_id,
        "visitDeptId": visit_dept_id,
        "opinion": opinion,
        "attachmentList": [],
        "outsideReply": None,
        "reply": ""
    }
    
    headers = {
        "Content-Type": "application/json",
        "x-access-token": token
    }
    
    try:
        response = requests.post(url, json=data, headers=headers, verify=False)
        response.raise_for_status()
        result = response.json()
        return {"success": True, "data": result}
    except Exception as e:
        return {"success": False, "message": str(e)}

result = select_file_and_token()

if result['cancelled']:
    sys.exit()

file_path = result['file_path']
token = result['token']

if not os.path.exists(file_path):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("错误", f"文件不存在：{file_path}")
    root.destroy()
    exit()

if result['save_token']:
    config = load_config()
    config['token'] = token
    save_config(config)

wb = load_workbook(file_path)
sheet = wb.active

# 添加结果列
headers = [cell.value for cell in sheet[1]]
if '回访结果' not in headers:
    sheet.cell(row=1, column=len(headers) + 1, value='回访结果')
    headers.append('回访结果')

success_count = 0
failed_count = 0

for row_idx in range(2, sheet.max_row + 1):
    issue_id = sheet.cell(row=row_idx, column=1).value
    circulation_id = sheet.cell(row=row_idx, column=2).value
    department_name = sheet.cell(row=row_idx, column=4).value
    
    if not issue_id or not circulation_id or not department_name:
        sheet.cell(row=row_idx, column=headers.index('回访结果') + 1, value='缺少必要数据')
        failed_count += 1
        continue
    
    result = add_return_visit(token, str(issue_id), str(circulation_id), str(department_name))
    
    if result['success']:
        sheet.cell(row=row_idx, column=headers.index('回访结果') + 1, value='成功')
        success_count += 1
    else:
        sheet.cell(row=row_idx, column=headers.index('回访结果') + 1, value=f'失败: {result["message"]}')
        failed_count += 1

try:
    wb.save(file_path)
except PermissionError:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("权限错误", "请关闭Excel文件，然后重试。")
    root.destroy()
    exit()

root = tk.Tk()
root.withdraw()
messagebox.showinfo("提示", f"回访部门添加完成！\n成功: {success_count} 条\n失败: {failed_count} 条")
root.destroy()
