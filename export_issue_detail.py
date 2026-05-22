# -*- coding: utf-8 -*-

# 导出纠纷详情
import os
import sys
import json
import datetime
import time

import pandas as pd
from openpyxl import load_workbook

import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

config_file = 'config.json'

def get_base_url():
    """从配置文件获取服务器基础URL"""
    config = load_config()
    host = config.get('server_host', 'your_server_ip_here')
    port = config.get('server_port', '39999')
    return f'https://{host}:{port}'

def get_api_url():
    return f"{get_base_url()}/event-center/issueDetail/exportIssueDetail"

DEFAULT_T = int(time.time())

def load_config():
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_config(config):
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def select_file_and_save_path():
    config = load_config()
    current_token = config.get('token', '')

    root = tk.Tk()
    root.title("导出纠纷详情")
    root.geometry("700x280")
    root.resizable(False, False)

    selected_file = tk.StringVar()
    save_path = tk.StringVar()
    token_var = tk.StringVar(value=current_token)
    save_token_var = tk.BooleanVar(value=False)

    result = {
        'file_path': '',
        'save_path': '',
        'token': current_token,
        'cancelled': True,
        'save_token': False
    }

    def browse_file():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            selected_file.set(file_path)

    def browse_save_path():
        folder_path = filedialog.askdirectory(title="选择保存位置")
        if folder_path:
            save_path.set(folder_path)

    def confirm():
        result['file_path'] = selected_file.get()
        result['save_path'] = save_path.get()
        result['token'] = token_var.get()
        result['cancelled'] = False
        result['save_token'] = save_token_var.get()
        root.destroy()

    def cancel():
        result['cancelled'] = True
        root.destroy()

    def on_closing():
        result['cancelled'] = True
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    file_frame = tk.Frame(root)
    file_frame.pack(pady=10, fill=tk.X, padx=20)
    tk.Label(file_frame, text="请选择Excel文件：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    entry_frame = tk.Frame(file_frame)
    entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(entry_frame, textvariable=selected_file, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(entry_frame, text="浏览...", command=browse_file, relief=tk.FLAT, bg="#f0f0f0", padx=10, pady=2).pack(side=tk.LEFT)

    save_frame = tk.Frame(root)
    save_frame.pack(pady=10, fill=tk.X, padx=20)
    tk.Label(save_frame, text="请选择保存位置：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    save_entry_frame = tk.Frame(save_frame)
    save_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(save_entry_frame, textvariable=save_path, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(save_entry_frame, text="浏览...", command=browse_save_path, relief=tk.FLAT, bg="#f0f0f0", padx=10, pady=2).pack(side=tk.LEFT)

    token_frame = tk.Frame(root)
    token_frame.pack(pady=10, fill=tk.X, padx=20)
    tk.Label(token_frame, text="用户令牌：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    token_entry_frame = tk.Frame(token_frame)
    token_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(token_entry_frame, textvariable=token_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Checkbutton(token_entry_frame, text="保存令牌", variable=save_token_var).pack(side=tk.LEFT, padx=10)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=15)
    tk.Button(button_frame, text="确定", command=confirm, width=12, bg="#4CAF50", fg="white", relief=tk.FLAT, padx=15, pady=5).pack(side=tk.LEFT, padx=10)
    tk.Button(button_frame, text="取消", command=cancel, width=12, relief=tk.FLAT, bg="#f0f0f0", padx=15, pady=5).pack(side=tk.LEFT, padx=10)

    root.bind('<Return>', lambda event: confirm())

    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    root.mainloop()
    return result

def export_issue_detail(token, issue_id, save_dir, event_name):
    url = f"{get_api_url()}?_t={DEFAULT_T}&id={issue_id}"

    headers = {
        "x-access-token": token
    }

    try:
        response = requests.get(url, headers=headers, verify=False, stream=True)

        if response.status_code == 401:
            return {
                "success": False,
                "message": "401无权限",
                "issue_id": issue_id,
                "unauthorized": True
            }

        response.raise_for_status()

        content_type = response.headers.get('Content-Type', '')

        if 'application/json' in content_type:
            error_data = response.json()
            return {
                "success": False,
                "message": error_data.get('message', '未知错误'),
                "issue_id": issue_id
            }

        if event_name:
            filename = f"关于{event_name}的事件详情.docx"
        else:
            filename = f"关于{issue_id}的事件详情.docx"

        save_path = os.path.join(save_dir, filename)

        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        return {
            "success": True,
            "filename": filename,
            "save_path": save_path,
            "issue_id": issue_id
        }

    except requests.exceptions.ConnectionError:
        return {
            "success": False,
            "message": "网络连接失败",
            "issue_id": issue_id
        }
    except Exception as e:
        return {
            "success": False,
            "message": str(e),
            "issue_id": issue_id
        }

def show_progress_bar(total, title="导出进度"):
    progress_window = tk.Toplevel()
    progress_window.title(title)
    progress_window.geometry("400x120")
    progress_window.resizable(False, False)
    progress_window.transient()

    progress_var = tk.DoubleVar()
    progress_bar = tk.ttk.Progressbar(
        progress_window,
        variable=progress_var,
        maximum=100,
        length=350,
        mode='determinate'
    )
    progress_bar.pack(pady=20, padx=20)

    status_label = tk.Label(progress_window, text="准备开始...", font=("微软雅黑", 10))
    status_label.pack(pady=10)

    progress_window.update_idletasks()
    width = progress_window.winfo_width()
    height = progress_window.winfo_height()
    x = (progress_window.winfo_screenwidth() // 2) - (width // 2)
    y = (progress_window.winfo_screenheight() // 2) - (height // 2)
    progress_window.geometry(f'{width}x{height}+{x}+{y}')

    def update_progress(current, total, status):
        percent = (current / total) * 100
        progress_var.set(percent)
        status_label.config(text=status)
        progress_window.update()

    def close():
        progress_window.destroy()

    return progress_window, update_progress, close
def main():
    result = select_file_and_save_path()

    if result['cancelled']:
        sys.exit()

    file_path = result['file_path']
    save_path = result['save_path']
    token = result['token']

    if not file_path:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("错误", "请选择Excel文件")
        root.destroy()
        sys.exit()

    if not save_path:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("错误", "请选择保存位置")
        root.destroy()
        sys.exit()

    if not os.path.exists(file_path):
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("错误", f"文件不存在：{file_path}")
        root.destroy()
        sys.exit()

    if not os.path.exists(save_path):
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("错误", f"保存位置不存在：{save_path}")
        root.destroy()
        sys.exit()

    if result['save_token']:
        config = load_config()
        config['token'] = token
        save_config(config)

    try:
        wb = load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("错误", f"无法打开Excel文件：{str(e)}")
        root.destroy()
        sys.exit()

    issue_data = []
    for row_idx in range(2, sheet.max_row + 1):
        issue_id = sheet.cell(row=row_idx, column=1).value
        event_name = sheet.cell(row=row_idx, column=3).value
        if issue_id is not None and issue_id != '':
            issue_id_str = str(issue_id).strip()
            if issue_id_str and not issue_id_str.lower() == 'id':
                event_name_str = str(event_name).strip() if event_name is not None else ''
                issue_data.append({
                    'id': issue_id_str,
                    'event_name': event_name_str
                })

    if not issue_data:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("错误", "Excel文件中没有找到有效的ID数据")
        root.destroy()
        sys.exit()

    total = len(issue_data)
    success_count = 0
    failed_count = 0
    failed_list = []
    unauthorized = False

    root = tk.Tk()
    root.withdraw()
    progress_window, update_progress, close_progress = show_progress_bar(total)

    try:
        for idx, data in enumerate(issue_data, 1):
            issue_id = data['id']
            event_name = data['event_name']
            update_progress(idx - 1, total, f"正在处理：{event_name or issue_id} ({idx}/{total})")
            export_result = export_issue_detail(token, issue_id, save_path, event_name)

            if export_result.get('unauthorized'):
                unauthorized = True
                break

            if export_result['success']:
                success_count += 1
                update_progress(idx, total, f"成功：{event_name or issue_id} ({success_count}/{total})")
            else:
                failed_count += 1
                failed_list.append(f"ID: {issue_id} - {export_result['message']}")
                update_progress(idx, total, f"失败：{event_name or issue_id} ({success_count}/{total})")

        update_progress(total, total, "任务完成！")
        time.sleep(0.5)
    finally:
        close_progress()

    if unauthorized:
        messagebox.showerror("无权限", "令牌已失效，请重新输入令牌！")
        config = load_config()
        config.pop('token', None)
        save_config(config)
        main()
        return

    if failed_count > 0:
        failed_info = "\n".join(failed_list[:10])
        if failed_count > 10:
            failed_info += f"\n... 还有 {failed_count - 10} 条失败记录"
        messagebox.showinfo(
            "导出完成",
            f"导出完成！\n成功: {success_count} 条\n失败: {failed_count} 条\n\n失败详情：\n{failed_info}"
        )
    else:
        messagebox.showinfo("导出完成", f"导出完成！\n成功导出: {success_count} 条")

    root.destroy()

if __name__ == "__main__":
    main()
