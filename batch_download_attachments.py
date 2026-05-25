# -*- coding: utf-8 -*-
"""
批量下载附件模块
从Excel表格中读取纠纷ID，批量下载附件
"""

import os
import sys
import json
import time
import requests
import urllib3
import pandas as pd
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock, Thread
import queue

import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 导入配置模块
from src.core.config import get_config

def get_app_path():
    """获取应用程序路径"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

def clean_filename(filename):
    """清理文件名中的非法字符"""
    illegal_chars = '\\/:*?"<>|'
    for char in illegal_chars:
        filename = filename.replace(char, '')
    filename = filename.replace('\n', '').replace('\r', '').replace('\t', '')
    filename = filename.strip()
    return filename

class DownloadSession:
    """下载会话管理"""
    _session = None
    
    @classmethod
    def get_session(cls):
        if cls._session is None:
            cls._session = requests.Session()
            cls._session.verify = False
            cls._session.trust_env = False
            adapter = requests.adapters.HTTPAdapter(
                pool_connections=10,
                pool_maxsize=20,
                max_retries=2
            )
            cls._session.mount('https://', adapter)
            cls._session.mount('http://', adapter)
        return cls._session

def download_file(url, token, save_path):
    """下载单个文件"""
    session = DownloadSession.get_session()
    try:
        headers = {'x-access-token': token} if token else {}
        response = session.get(url, headers=headers, timeout=60)
        response.raise_for_status()
        
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return True, None
    except requests.exceptions.RequestException as e:
        error_msg = f"网络请求失败: {str(e)}"
        if '401' in str(e) or 'Unauthorized' in str(e):
            error_msg = "权限不足，请检查令牌是否有效"
        elif 'Connection refused' in str(e):
            error_msg = "无法连接到服务器，请检查网络"
        elif 'timeout' in str(e).lower():
            error_msg = "请求超时，请稍后重试"
        return False, error_msg
    except Exception as e:
        return False, f"保存文件失败: {str(e)}"

def get_issue_detail(token, issue_id):
    """获取纠纷详情"""
    session = DownloadSession.get_session()
    config = get_config()
    base_url = config.get_base_url()
    api_url = f"{base_url}/event-center/issueDetail/getIssueDetail?id={issue_id}&_t={int(time.time())}"
    headers = {'x-access-token': token} if token else {}
    
    try:
        response = session.get(api_url, headers=headers, timeout=30)
        response.raise_for_status()
        
        try:
            return response.json(), None
        except json.JSONDecodeError:
            return None, "服务器返回无效数据"
    except requests.exceptions.RequestException as e:
        error_msg = f"API请求失败: {str(e)}"
        if '401' in str(e) or 'Unauthorized' in str(e):
            error_msg = "权限不足，令牌无效或已过期"
        elif 'Connection refused' in str(e):
            error_msg = "无法连接到服务器，请检查网络连接"
        elif 'SSLError' in str(e):
            error_msg = "SSL连接错误，请检查网络环境"
        elif 'timeout' in str(e).lower():
            error_msg = "请求超时，请稍后重试"
        return None, error_msg
    except Exception as e:
        return None, f"未知错误: {str(e)}"

def select_file_and_save_path(token=''):
    """选择文件和保存路径对话框"""
    config = get_config()
    current_token = token if token else config.get_token()

    root = tk.Tk()
    root.title("批量下载事件附件")
    root.geometry("700x320")
    root.resizable(False, False)
    root.attributes('-topmost', True)

    selected_file = tk.StringVar()
    save_path = tk.StringVar()
    token_var = tk.StringVar(value=current_token)
    save_token_var = tk.BooleanVar(value=False)

    result = {
        'file_path': '',
        'save_path': '',
        'token': current_token,
        'cancelled': True,
        'save_token': False
    }

    def browse_file():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
            parent=root
        )
        if file_path:
            selected_file.set(file_path)

    def browse_save_path():
        folder_path = filedialog.askdirectory(title="选择保存位置", parent=root)
        if folder_path:
            save_path.set(folder_path)

    def confirm():
        result['file_path'] = selected_file.get()
        result['save_path'] = save_path.get()
        result['token'] = token_var.get()
        result['cancelled'] = False
        result['save_token'] = save_token_var.get()
        root.destroy()

    def cancel():
        result['cancelled'] = True
        root.destroy()

    def on_closing():
        result['cancelled'] = True
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    file_frame = tk.Frame(root)
    file_frame.pack(pady=10, fill=tk.X, padx=20)
    tk.Label(file_frame, text="请选择Excel文件：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    entry_frame = tk.Frame(file_frame)
    entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(entry_frame, textvariable=selected_file, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(entry_frame, text="浏览...", command=browse_file, relief=tk.FLAT, bg="#f0f0f0", padx=10, pady=2).pack(side=tk.LEFT)

    save_frame = tk.Frame(root)
    save_frame.pack(pady=10, fill=tk.X, padx=20)
    tk.Label(save_frame, text="请选择保存位置：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    save_entry_frame = tk.Frame(save_frame)
    save_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(save_entry_frame, textvariable=save_path, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(save_entry_frame, text="浏览...", command=browse_save_path, relief=tk.FLAT, bg="#f0f0f0", padx=10, pady=2).pack(side=tk.LEFT)

    token_frame = tk.Frame(root)
    token_frame.pack(pady=10, fill=tk.X, padx=20)
    tk.Label(token_frame, text="用户令牌：", font=("微软雅黑", 10), width=15, anchor=tk.W).pack(side=tk.LEFT, padx=5)
    token_entry_frame = tk.Frame(token_frame)
    token_entry_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Entry(token_entry_frame, textvariable=token_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Checkbutton(token_entry_frame, text="保存令牌", variable=save_token_var).pack(side=tk.LEFT, padx=10)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=15)
    tk.Button(button_frame, text="确定", command=confirm, width=12, bg="#4CAF50", fg="white", relief=tk.FLAT, padx=15, pady=5).pack(side=tk.LEFT, padx=10)
    tk.Button(button_frame, text="取消", command=cancel, width=12, relief=tk.FLAT, bg="#f0f0f0", padx=15, pady=5).pack(side=tk.LEFT, padx=10)

    root.bind('<Return>', lambda event: confirm())

    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    root.mainloop()
    return result

class ProgressWindow:
    """进度窗口"""
    def __init__(self, total, title="下载进度"):
        self.progress_window = tk.Toplevel()
        self.progress_window.title(title)
        self.progress_window.geometry("550x180")
        self.progress_window.resizable(False, False)
        self.progress_window.transient()

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.progress_window,
            variable=self.progress_var,
            maximum=100,
            length=500,
            mode='determinate'
        )
        self.progress_bar.pack(pady=20, padx=20)

        self.status_label = tk.Label(self.progress_window, text="准备开始...", font=("微软雅黑", 10))
        self.status_label.pack(pady=5)

        self.detail_label = tk.Label(self.progress_window, text="", font=("微软雅黑", 9), fg="#666666")
        self.detail_label.pack(pady=5)

        self.error_label = tk.Label(self.progress_window, text="", font=("微软雅黑", 9), fg="#ff4444")
        self.error_label.pack(pady=5)

        self.progress_window.update_idletasks()
        width = self.progress_window.winfo_width()
        height = self.progress_window.winfo_height()
        x = (self.progress_window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.progress_window.winfo_screenheight() // 2) - (height // 2)
        self.progress_window.geometry(f'{width}x{height}+{x}+{y}')

        self.last_update_time = 0
        self.update_queue = queue.Queue()
        self._schedule_update()

    def _schedule_update(self):
        try:
            while True:
                args = self.update_queue.get_nowait()
                self._do_update(*args)
        except queue.Empty:
            pass
        self.progress_window.after(50, self._schedule_update)

    def _do_update(self, current, total, status, detail="", error=""):
        now = time.time()
        if now - self.last_update_time < 0.05:
            return
        self.last_update_time = now
        
        percent = (current / total) * 100
        self.progress_var.set(percent)
        self.status_label.config(text=status)
        self.detail_label.config(text=detail)
        self.error_label.config(text=error)
        self.progress_window.update_idletasks()

    def update_progress(self, current, total, status, detail="", error=""):
        self.update_queue.put((current, total, status, detail, error))

    def close(self):
        self.progress_window.destroy()

def show_message(title, message, icon='info'):
    """显示消息对话框"""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    if icon == 'error':
        messagebox.showerror(title, message)
    elif icon == 'warning':
        messagebox.showwarning(title, message)
    else:
        messagebox.showinfo(title, message)
    root.destroy()

def show_confirm(title, message):
    """显示确认对话框"""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    result = messagebox.askyesno(title, message)
    root.destroy()
    return result

def download_attachments_parallel(token, attachments, folder_path):
    """并行下载附件"""
    failed_list = []
    success_count = 0
    lock = Lock()
    
    def download_task(attachment):
        nonlocal success_count
        full_url = attachment.get('fullUrl', '')
        file_name = attachment.get('fileName', '')
        
        if not full_url or not file_name:
            return
        
        file_name_clean = clean_filename(file_name)
        save_file_path = os.path.join(folder_path, file_name_clean)
        
        success, err_msg = download_file(full_url, token, save_file_path)
        with lock:
            if success:
                success_count += 1
            else:
                failed_list.append(f"附件下载失败: {file_name} - {err_msg}")
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(download_task, attachment) for attachment in attachments]
        for future in as_completed(futures):
            future.result()
    
    return success_count, failed_list

def download_worker(token, issue_data, save_path, update_progress_func, result_queue):
    """下载工作线程"""
    total = len(issue_data)
    success_count = 0
    failed_count = 0
    total_attachments = 0
    failed_list = []
    unauthorized = False
    last_error = ""

    try:
        for idx, data in enumerate(issue_data, 1):
            issue_id = data['id']
            event_name = data['event_name']
            folder_prefix = data['folder_prefix']
            
            update_progress_func(idx - 1, total, f"正在处理事件 {idx}/{total}", f"ID: {issue_id} | {event_name}", "")

            detail_data, error = get_issue_detail(token, issue_id)
            
            if error:
                if '权限不足' in error or '401' in error or '令牌' in error:
                    unauthorized = True
                    last_error = error
                    break
                failed_count += 1
                failed_list.append(f"ID: {issue_id} - {error}")
                continue

            if detail_data is None:
                failed_count += 1
                failed_list.append(f"ID: {issue_id} - 获取数据失败")
                continue

            if detail_data.get('code') == 401:
                unauthorized = True
                last_error = "令牌无效或已过期，请重新输入"
                break

            if detail_data.get('success') is not None and not detail_data['success']:
                msg = detail_data.get('message', '操作失败')
                if '权限' in msg or '401' in msg:
                    unauthorized = True
                    last_error = msg
                    break
                failed_count += 1
                failed_list.append(f"ID: {issue_id} - {msg}")
                continue

            result_data = detail_data.get('result', {})
            
            folder_name = clean_filename(f"{folder_prefix}({event_name})" if folder_prefix and event_name else (folder_prefix if folder_prefix else str(issue_id)))
            folder_path = os.path.join(save_path, folder_name)
            
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                update_progress_func(idx, total, f"处理事件 {idx}/{total}", f"创建文件夹: {folder_name}", "")

            operation_logs = result_data.get('operationLogs', [])
            all_attachments = []
            
            for log in operation_logs:
                process_dept = log.get('processDept', '')
                # 通用匹配：包含"街道"的部门
                if process_dept and '街道' in process_dept:
                    attachments = log.get('attachmentList', [])
                    all_attachments.extend(attachments)

            if all_attachments:
                update_progress_func(idx, total, f"处理事件 {idx}/{total}", f"下载附件中... ({len(all_attachments)}个)", "")
                
                event_success, event_failed = download_attachments_parallel(token, all_attachments, folder_path)
                total_attachments += event_success
                failed_count += len(event_failed)
                failed_list.extend([f"ID: {issue_id} - {f}" for f in event_failed])
            else:
                update_progress_func(idx, total, f"处理事件 {idx}/{total}", f"ID: {issue_id} | 无附件", "")

            success_count += 1
            update_progress_func(idx, total, f"处理完成 {idx}/{total}", f"ID: {issue_id} | 下载附件: {len(all_attachments)}个", "")

        if not unauthorized:
            update_progress_func(total, total, "任务完成！", f"共下载 {total_attachments} 个附件", "")
            time.sleep(0.5)
    finally:
        DownloadSession._session = None

    result_queue.put({
        'success_count': success_count,
        'failed_count': failed_count,
        'total_attachments': total_attachments,
        'failed_list': failed_list,
        'unauthorized': unauthorized,
        'last_error': last_error
    })

def main(token=''):
    """主函数"""
    result = select_file_and_save_path(token)

    if result['cancelled']:
        sys.exit()

    file_path = result['file_path']
    save_path = result['save_path']
    token = result['token']

    if not file_path:
        show_message("错误", "请选择Excel文件", 'error')
        sys.exit()

    if not save_path:
        show_message("错误", "请选择保存位置", 'error')
        sys.exit()

    if not os.path.exists(file_path):
        show_message("错误", f"文件不存在：{file_path}", 'error')
        sys.exit()

    if not os.path.exists(save_path):
        show_message("错误", f"保存位置不存在：{save_path}", 'error')
        sys.exit()

    if not token.strip():
        choice = show_confirm("提示", "未输入用户令牌，可能无法正常下载附件。是否继续？")
        if not choice:
            sys.exit()

    if result['save_token']:
        config = get_config()
        config.set('token', token)
        config.save()

    try:
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
    except Exception as e:
        show_message("错误", f"无法打开Excel文件：{str(e)}", 'error')
        sys.exit()

    issue_data = []
    for row_idx in range(2, sheet.max_row + 1):
        issue_id = sheet.cell(row=row_idx, column=1).value
        event_name = sheet.cell(row=row_idx, column=3).value if sheet.max_column >= 3 else None
        folder_prefix = sheet.cell(row=row_idx, column=4).value if sheet.max_column >= 4 else None
        if issue_id is not None and issue_id != '':
            issue_id_str = str(issue_id).strip()
            if issue_id_str and not issue_id_str.lower() == 'id':
                event_name_str = str(event_name).strip() if event_name is not None else ''
                folder_prefix_str = str(folder_prefix).strip() if folder_prefix is not None else ''
                issue_data.append({
                    'id': issue_id_str,
                    'event_name': event_name_str,
                    'folder_prefix': folder_prefix_str,
                    'row_idx': row_idx
                })

    wb.close()

    if not issue_data:
        show_message("错误", "Excel文件中没有找到有效的ID数据", 'error')
        sys.exit()

    total = len(issue_data)

    root = tk.Tk()
    root.withdraw()
    progress_window = ProgressWindow(total)

    result_queue = queue.Queue()
    
    worker_thread = Thread(
        target=download_worker,
        args=(token, issue_data, save_path, progress_window.update_progress, result_queue),
        daemon=True
    )
    worker_thread.start()

    def check_worker():
        if worker_thread.is_alive():
            root.after(100, check_worker)
        else:
            try:
                result_data = result_queue.get_nowait()
                progress_window.close()
                root.destroy()

                if result_data['unauthorized']:
                    retry = show_confirm("权限不足", f"{result_data['last_error']}\n是否重新输入令牌并重试？")
                    if retry:
                        config = get_config()
                        config.set('token', '')
                        config.save()
                        main()
                        return
                    else:
                        show_message("提示", "操作已取消", 'info')
                        return

                if result_data['failed_count'] > 0:
                    failed_info = "\n".join(result_data['failed_list'][:10])
                    if result_data['failed_count'] > 10:
                        failed_info += f"\n... 还有 {result_data['failed_count'] - 10} 条失败记录"
                    show_message(
                        "下载完成",
                        f"下载完成！\n成功处理: {result_data['success_count']} 条事件\n共下载: {result_data['total_attachments']} 个附件\n失败: {result_data['failed_count']} 条\n\n失败详情：\n{failed_info}",
                        'info'
                    )
                else:
                    show_message("下载完成", f"下载完成！\n成功处理: {result_data['success_count']} 条事件\n共下载: {result_data['total_attachments']} 个附件", 'info')
            except queue.Empty:
                pass

    root.after(100, check_worker)
    root.mainloop()

if __name__ == "__main__":
    main()
